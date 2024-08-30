# -*- coding: utf-8 -*-
"""
Created on Mon Jun 17 10:38:35 2024

@author: vi.nt
"""

import pandas as pd
from openpyxl import load_workbook
import json
import requests
from datetime import datetime
import curlify

# Step 0
def ttl_api_log_in(operatorID, password, channelID):
    url = 'http://172.25.11.16:7666/operatorLogin'
    #url = 'http://172.25.11.17:7666/operatorLogin'  # enhance performance
    #url = 'http://172.25.15.10:7666/operatorLogin'  # UAT
    
    param_json = json.dumps({
        'operatorID': operatorID,
        'password': password,
        'channelID': channelID})
    
    headers = {
      'Content-Type': 'application/json',
      'Cookie': 'vertx-web.session=5390a81f66b42bbb25ca4b473a911ec0'
      }

    
    resp = requests.post(url, headers=headers, data=param_json)
    #print(resp.text)
    resp_dict = json.loads(resp.text)
    token = resp_dict['tokenID']
    
    if resp_dict['errorCode'] == 'OLS0000':
        token = resp_dict['tokenID']
        exp = resp_dict['exp']
    
    return token, exp

# Step 1
def load_excel_data(path):
    
    in_sheet_name = 'input'
    df_input = pd.read_excel(path, sheet_name = in_sheet_name)

    return df_input

# function Step 2
def get_sub_trading_data_today_equity(token, subAccountID, tradingAccSeq, 
                                      clientID, operatorID, 
                                      start = 0, limit = 1000, 
                                      loop_count = 10000):
    print(subAccountID)
    # init
    df_all_list = []
    url = 'http://172.25.11.16:7666/services/eqt/enquiryorder'
    #url = 'http://172.25.11.17:7666/services/eqt/enquiryorder' # tăng performance
    start = start
    limit= limit
    headers = {
      'Content-Type': 'application/json'
      #'Cookie': 'vertx-web.session=5390a81f66b42bbb25ca4b473a911ec0'
    }
    
    count = 0
    while True:
        if count >= loop_count:
            break
   
        param_json =  json.dumps({
                  "subAccountID": subAccountID,
                  "clientID": clientID,
                  "tradingAccSeq": tradingAccSeq,
                  "token": token,
                  "operatorID": operatorID,
                  "start": start,
                  "limit": limit,
                  "mvStatus": "MATCHED",
                  "mvInstrumentID": "ALL"
                  })
        #print(param_json)
        resp = requests.post(url, headers=headers, data=param_json)
        #print(curlify.to_curl(resp.request))

        
        resp_dict = json.loads(resp.text)
        
        if resp_dict['errorCode'] == 'OLS0000':
            if resp_dict['mvOrderBeanList']:
                df = pd.DataFrame(resp_dict['mvOrderBeanList'])
                df['mvFilledQty'] = df['mvFilledQty'].str.replace(',','').astype(int)
                df = df[df['mvFilledQty']>0]
                
                df_all_list.append(df)
                
            if resp_dict['mvOrderBeanList'] == None:
                print(resp_dict['mvMessage'])
               
                # to save start variable 
                break
            
            if len(resp_dict['mvOrderBeanList']) < limit:
                #print('No more data to query')
                start += len(resp_dict['mvOrderBeanList'])
                #print('start:', start)
                # save start var
                break
        else:
            print(resp_dict['errorCode'])
            print(resp_dict['errorMessage'])
           
            # save start var
            break
        
        count += 1
        
        start += limit
        print('start:', start)
    if df_all_list:
        df = pd.concat(df_all_list)
    else:
        print('Dataframe None')
        df = pd.DataFrame()
    return df

# function Step 3
def validate_token():
    pass

def get_sub_trading_data_today_future(token, subAccountID, tradingAccSeq, clientID, operatorID, start = 0, limit = 1000):
    print(subAccountID)
    # init
    df_all_list = []
    url = 'http://172.25.11.16:7666/services/fno/orderenquiry'
    #url = 'http://172.25.11.17:7666/services/fno/orderenquiry' # enhance performance
    start = start
    limit= limit
    headers = {
      'Content-Type': 'application/json',
      'Cookie': 'vertx-web.session=5390a81f66b42bbb25ca4b473a911ec0'
    }
    
    
    param_json =  json.dumps({
              "subAccountID": subAccountID,
              "token": token,
              "clientID": clientID,
              "tradingAccSeq": tradingAccSeq,
              "operatorID": operatorID,
              "status": 'FILLED',
              })

    resp = requests.post(url, headers=headers, data=param_json)
    
    resp_dict = json.loads(resp.text)
    
    if resp_dict['errorCode'] == 'OLS0000':
        if resp_dict['orderEnquiryInfoList']:
            df = pd.DataFrame(resp_dict['orderEnquiryInfoList'])
            df['qty'] = df['qty'].astype(int)
            df = df[df['qty']>0]
            
            df_all_list.append(df)
            
        if resp_dict['orderEnquiryInfoList'] == None:
            print(resp_dict['mvMessage'])
            
    else:
        print(resp_dict['errorCode'])
        print(resp_dict['errorMessage'])
               
    if df_all_list:
        df = pd.concat(df_all_list)
    else:
        #print('Dataframe None')
        df = pd.DataFrame()
    return df, resp_dict

# Step 4
def calculate_total_buy_sell(df_input, df_equity_list, df_future):
    df_new = pd.DataFrame(columns = ['ticker', 'buy', 'buyAmount', 
                                     'sell', 'sellAmount', 'subAccount'])
    for i, value in enumerate(df_equity_list):
        if value.empty:
            continue
        else:
            df_equity = value
        df_output = pd.DataFrame(columns = ['ticker', 'buy', 'buyAmount', 
                                            'sell', 'sellAmount', 'subAccount'])
        
        df_output.ticker = df_equity[['mvStockID', 
                                      'mvFilledQty']].groupby(['mvStockID']).sum().index
        df_equity[['mvGrossAmt']] = df_equity[['mvGrossAmt']].astype(float)
        df_output.buy = df_equity[['mvStockID', 'mvFilledQty']][df_equity.mvBS == 'B'].groupby(['mvStockID']).sum().join(df_output[['ticker']].set_index('ticker'), how='outer').mvFilledQty.values
        df_output.buyAmount = df_equity[['mvStockID', 'mvGrossAmt']][df_equity.mvBS == 'B'].groupby(['mvStockID']).sum().join(df_output[['ticker']].set_index('ticker'), how='outer').mvGrossAmt.values
        df_output.sell = df_equity[['mvStockID', 'mvFilledQty']][df_equity.mvBS == 'S'].groupby(['mvStockID']).sum().join(df_output[['ticker']].set_index('ticker'), how='outer').mvFilledQty.values
        df_output.sellAmount = df_equity[['mvStockID', 'mvGrossAmt']][df_equity.mvBS == 'S'].groupby(['mvStockID']).sum().join(df_output[['ticker']].set_index('ticker'), how='outer').mvGrossAmt.values
        df_output.subAccount = df_input.loc[i,'subAccount']
        df_new = pd.concat([df_new, df_output], axis=0)

    df_future = df_future[df_future.filled > 0]
    df_out_future = pd.DataFrame(columns = ['ticker', 'bs', 'average', 'filled', 'amount'])
    for i, value in enumerate(df_future.orderInfo):
        df_out_future.loc[len(df_out_future.index)] = [value.get('seriesID'), value.get('bs'), df_future.average[i], df_future.filled[i], df_future.average[i] * df_future.filled[i] * 100000]    
    # print(df_output)
    df_output_future = pd.DataFrame(columns = df_output.columns)
    df_output_future.ticker = df_out_future[['ticker', 'amount']].groupby(['ticker']).sum().index
    df_output_future.buy = df_out_future[['ticker', 'filled']][df_out_future.bs == 'B'].groupby(['ticker']).sum().join(df_out_future[['ticker', 'amount']].groupby(['ticker']).sum(), how='outer').filled.values
    df_output_future.buyAmount = df_out_future[['ticker', 'amount']][df_out_future.bs == 'B'].groupby(['ticker']).sum().join(df_out_future[['ticker', 'filled']].groupby(['ticker']).sum(), how='outer').amount.values
    # print(df_output_future)

    df_output_future.sell = df_out_future[['ticker', 'filled']][df_out_future.bs == 'S'].groupby(['ticker']).sum().join(df_out_future[['ticker', 'amount']].groupby(['ticker']).sum(), how='outer').filled.values
    df_output_future.sellAmount = df_out_future[['ticker', 'amount']][df_out_future.bs == 'S'].groupby(['ticker']).sum().join(df_out_future[['ticker', 'filled']].groupby(['ticker']).sum(), how='outer').amount.values
    df_output_future.subAccount = df_input.loc[len(df_input)-1,'subAccount']
    df_new = pd.concat([df_new, df_output_future], axis=0)
    df_output1 = pd.DataFrame(columns = ['Time', 'Buy', 'Sell', 'Delta'])
    df_output1.loc[0, 'Time'] = pd.to_datetime(str(datetime.now())+'+00:00').strftime("%d/%m/%Y %H:%M:%S")
    # df_output1.loc[0, 'Time'] = pd.to_datetime(str(datetime.now())+'+00:00').tz_convert('Etc/GMT-7').strftime("%d/%m/%Y %H:%M:%S")
    df_output1.loc[0, 'Buy'] = df_new.buyAmount.sum()
    df_output1.loc[0, 'Sell'] = df_new.sellAmount.sum()
    df_output1.loc[0, 'Delta'] = df_output1.loc[0, 'Buy'] - df_output1.loc[0, 'Sell']
    return df_new, df_output1

def calculate_buy_sell_not_include_futures(df_buy_sell, df_delta):
    df = df_buy_sell[df_buy_sell.subAccount != 'ECB5693D1']
    buy = df.buyAmount.sum()
    sell = df.sellAmount.sum()
    df_delta['Buy (not include futures)'] = buy
    df_delta['Sell (not include futures)'] = sell
    df_delta['Delta (not include futures)'] = buy - sell
    return df_delta
# Step 5
def save_df_to_file(df, df1, path, sheet_name, sheet_name1):
    wb = load_workbook(path)
    ws = wb[sheet_name]
    ind = 1
    while True:
        if not pd.isna(ws.cell(row= ind, column= 1).value): ind += 1
        else: break
    for i, col in enumerate(df):
        for j, row in enumerate(df[col]):
            ws.cell(row= j + ind, column= i + 1).value = df.iloc[j,i]  
    wb.save(path)
    
    # wb = load_workbook(path)
    # ws = wb[sheet_name1]
    # for i, col in enumerate(df1):
    #     for j, row in enumerate(df1[col]):
    #         ws.cell(row= j + 2, column= i + 1).value = df1.iloc[j,i] 
    # wb.save(path)
    
    with pd.ExcelWriter(path, engine="openpyxl", mode='a', if_sheet_exists='replace') as writer:  
        df1.to_excel(writer, sheet_name=sheet_name1, index=False)
        
############### MAIN ################################################################

def main():
    operatorID = 'UYEN.LH'

    # Step 0.0: load data
    
    
    # Step 1: load current data, start_array of each ticker
    path = 'delta_data_1.3.xlsx'
    df_input = load_excel_data(path)
    
    #Step 0: log in
    try:
        data = json.load(open('token.json'))
        token = data['token']
        expired_dt = datetime.fromtimestamp(int(data['exp'])) 
        
        # get_sub_trading_data_today_equity(token, df_input.loc[0,'subAccount'], int(df_input.loc[0,'tradeSeq']), 
        #                                           df_input.loc[0,'clientID'], operatorID, 
        #                                           start = int(df_input.loc[0,'start']), limit = 1, loop_count= 1)
        
        if datetime.now() < expired_dt:
            print('token not expired')
        else:
            raise Exception("token expired")

    except:
        print('login to get new token')
        #token, exp = ttl_api_log_in(operatorID, 'abcd12345!', 'SS')
        token, exp = ttl_api_log_in(operatorID, 'Wsnewera2024', 'SS')
        
        data = {'token':  token,
                'exp': exp} 
        with open("token.json", "w+") as f:
            json.dump(data, f)
            
        
    # token = ttl_api_log_in(operatorID, 'data12345!', 'SS')
    
    # data = {'token':  token}
    # with open("token.json", "w+") as f:
    #     json.dump(data, f)
    
    # Step 2: get api trade data for equity (do until no more data)
    df_equity_list = []
    for i in range(len(df_input)-1):
        df_equity_list.append(get_sub_trading_data_today_equity(token, df_input.loc[i,'subAccount'], int(df_input.loc[i,'tradeSeq']), 
                                                  df_input.loc[i,'clientID'], operatorID, 
                                                  start = int(df_input.loc[i,'start']), limit = 1000))
    
    # # Step 3: get api trade data for Future (do until no more data)
    df_future, resp_dict = get_sub_trading_data_today_future(token, df_input.loc[len(df_input)-1,'subAccount'], int(df_input.loc[len(df_input)-1,'tradeSeq']), 
                                                  df_input.loc[len(df_input)-1,'clientID'], operatorID, 
                                                  start = int(df_input.loc[len(df_input)-1,'start']), limit = 1000)
    
    # Step 4: update total buy/sell data to current data (read from Excel in step 1)
    df_buy_sell, df_output = calculate_total_buy_sell(df_input, df_equity_list, df_future)
    df_delta = calculate_buy_sell_not_include_futures(df_buy_sell, df_output)
    # Step 5: save all to file in Step 1
    save_df_to_file(df_delta, df_buy_sell, path, 'sumBuySell', 'output_tradeData')